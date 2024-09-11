from image_processing_engine import color_matches
import pyautogui
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from PIL import Image
import numpy as np
import cv2

# Path to your Accountant Excel file
excel_file = r'Accountant.xlsx'  # Update with your path


def save_matrix_to_file(image_matrix, target_color, tolerance,
                        filename="pixel_matrix.txt", raw_filename="raw_matrix.txt"):
    # Get the current directory path
    current_directory = os.getcwd()
    # Define the path to the Auto_Lvler_Images directory within the current directory
    save_directory = os.path.join(current_directory, "Auto_Lvler_Images")

    # Construct the full file paths
    file_path = os.path.join(save_directory, filename)
    raw_file_path = os.path.join(save_directory, raw_filename)

    with open(file_path, 'w') as file:
        for row in image_matrix:
            for pixel in row:
                # Check if the current pixel matches the target color within the given tolerance
                if color_matches(pixel, target_color, tolerance):
                    # Represent matching pixels distinctly, e.g., with a special marker
                    file.write('X ')
                else:
                    # Optionally, convert non-matching pixels to a simplified representation
                    file.write('. ')
            file.write('\n')  # New line at the end of each row
    # Save raw RGB values
    with open(raw_file_path, 'w') as raw_file:
        for row in image_matrix:
            # Join RGB values for each pixel in the row, formatted as (R, G, B)
            row_str = ' '.join(f'({pixel[0]}, {pixel[1]}, {pixel[2]})' for pixel in row)
            raw_file.write(row_str + '\n')


def screenshot_saving_to_file(screenshot, ss_name):
    # Check if the screenshot is a NumPy array and determine the correct conversion
    if isinstance(screenshot, np.ndarray):
        # Handle BGRA (4 channels) to RGBA conversion for PIL compatibility
        if screenshot.shape[2] == 4:
            # Convert BGRA to RGBA
            screenshot = cv2.cvtColor(screenshot, cv2.COLOR_BGRA2RGBA)
        elif screenshot.shape[2] == 3:
            # Convert BGR to RGB for 3 channel images
            screenshot = cv2.cvtColor(screenshot, cv2.COLOR_BGR2RGB)
        # Convert the corrected NumPy array to a PIL Image object
        image = Image.fromarray(screenshot)
    elif isinstance(screenshot, Image.Image):
        # If screenshot is already a PIL Image object, use it directly
        print("Saving image directly as PIL object - ")
        image = screenshot
    else:
        raise TypeError("Unsupported image type")

    # Define the save directory
    current_directory = os.getcwd()
    save_directory = os.path.join(current_directory, "Auto_Lvler_Images")
    if not os.path.exists(save_directory):
        os.makedirs(save_directory)

    # Define the full path for the screenshot file
    screenshot_file_path = os.path.join(save_directory, ss_name)

    # Save the image
    image.save(screenshot_file_path)


def save_to_file_searched_image(search_area_size=100, save_path=r"Auto_Lvler_Images\Searched_image.jpg"):
    # Get the screen size
    screen_width, screen_height = pyautogui.size()

    # Calculate the center of the screen
    center_x, center_y = screen_width // 2, screen_height // 2

    # Define the region to capture: (left, top, width, height)
    region = (
        center_x - search_area_size,
        center_y - search_area_size,
        search_area_size * 2,
        search_area_size * 2
    )

    # Take a screenshot of the defined region
    screenshot = pyautogui.screenshot(region=region)

    # Save the screenshot
    screenshot.save(save_path)
    print(f"Screenshot saved to {save_path}")


def read_excel_counters(chars):
    try:
        # Attempt to read the existing Excel file
        if os.path.exists(excel_file) and os.path.getsize(excel_file) > 0:
            df = pd.read_excel(excel_file, index_col=0)
        else:
            # If the file doesn't exist, create an empty DataFrame
            df = pd.DataFrame(index=['Meteors', 'Successes'])

        # Add new columns for any new characters
        for character in chars:
            if character not in df.columns:
                df[character] = [0, 0]  # Assuming two rows: 'Meteors', 'Successes'

        return df

    except Exception as e:
        print(f"Error reading or initializing Excel file: {e}")
        return pd.DataFrame(columns=chars, index=['Meteors', 'Successes'])


def save_excel_with_formatting(dfx, counter_meteors, character):
    # Determine the columns that contain character data and should be summed
    character_columns = [col for col in dfx.columns if col not in ['Total']]
    # Increment the counter
    if character in dfx.columns:
        print("Updating & Saving Excel File... ")
        dfx.at['Meteors', character] = dfx.at['Meteors', character] + counter_meteors
    else:
        print("Character or Counter Type not found")

    # Add a 'Total' column that sums the values of the character columns for each row
    dfx['Total'] = dfx[character_columns].sum(axis=1)

    # Save the DataFrame to an Excel file
    dfx.to_excel(excel_file, engine='openpyxl')

    apply_formatting_to_excel(excel_file)  # Call formatting function


def apply_formatting_to_excel(excel_file_path):
    book = load_workbook(excel_file_path)
    sheet = book.active
    bold_font = Font(size=12, bold=True)
    meteor_font = Font(size=14, color="7030A0", bold=True)  # Purple color
    success_font = Font(size=14, color="00FF00", bold=True)  # Green color
    # loop_start_time = time.time()

    for cell in sheet[1]:  # Assuming row 1 contains the headers
        cell.font = bold_font

    for row in sheet.iter_rows(min_row=2, max_row=3, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.value == 'Meteors':
                cell.font = meteor_font
            elif cell.value == 'Successes':
                cell.font = success_font
            elif cell.coordinate.startswith('A'):
                cell.font = bold_font

    # Auto-adjust columns' width
    for column_cells in sheet.columns:
        max_length = 0
        column = [cell for cell in column_cells if cell.value]  # Only consider cells with content
        for cell in column:
            try:
                cell_value = str(cell.value)  # Convert cell value to string
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except (TypeError, AttributeError):
                pass
        adjusted_width = (max_length + 4)  # Adding a little extra space
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        book.save(excel_file_path)
